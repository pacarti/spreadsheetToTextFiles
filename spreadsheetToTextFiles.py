def column_max_row(sheet, column_name):
    return max((c.row for c in sheet[column_name] if c.value is not None))

import os, openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

os.chdir(os.path.dirname(os.path.abspath(__file__)))

wb = openpyxl.load_workbook('resultSpreadsheetFromTxtFiles.xlsx', data_only=True)

sheet = wb.active


# Create a loop from below to make one file per column: 

path = Path('textFiles')

if path.exists() == False:
    path.mkdir()

for j in range(1, sheet.max_column + 1):

    textFileName = 'textFile' + get_column_letter(j) + '.txt'
    textFile = open(path/textFileName, 'w')

    for i in range(1, column_max_row(sheet, get_column_letter(j)) + 1):

        cellValue = sheet[get_column_letter(j) + str(i)].value
        print(cellValue) # To see what will be written into the spreadsheet.

        if cellValue is not None:
            textFile.write(cellValue + '\n')
        else:
            textFile.write('\n')    
        

    textFile.close()